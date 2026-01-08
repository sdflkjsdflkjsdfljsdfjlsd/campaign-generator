import os
import random
import streamlit as st
from openpyxl import load_workbook


# --- 1. DEFINITIONS ---

def parse_campaign_data():
    """Parses 'campaign_data.xlsx' for missions, intensity, duration, and OpFor."""
    # For Streamlit deployment, it's often safer to use simple paths
    file_path = 'campaign_data.xlsx'

    # Fallback to local script dir if running locally and file isn't in root
    if not os.path.exists(file_path):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(script_dir, 'campaign_data.xlsx')

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    campaign_data = {}
    for col in ws.iter_cols(values_only=True):
        campaign_name = col[0]
        if campaign_name is None:
            continue

        campaign_name = str(campaign_name).strip()
        campaign_dict = {}
        current_key = None

        for cell_value in col[1:]:
            if cell_value is None:
                current_key = None
                continue

            # Identify the start of a new section (Missions, Intensity, Duration, OpFor)
            clean_val = str(cell_value).strip().lower()
            if clean_val in ["missions", "intensity", "duration", "opfor"]:
                current_key = clean_val
                campaign_dict[current_key] = []
            elif current_key:
                val = cell_value
                if current_key == "duration" and isinstance(val, (int, float)):
                    val = int(val)
                elif current_key == "opfor" and isinstance(val, (int, float)):
                    val = float(val)
                campaign_dict[current_key].append(val)

        campaign_data[campaign_name] = campaign_dict
    return campaign_data


def parse_intensity_data():
    file_path = 'Intensity calcs.xlsx'
    if not os.path.exists(file_path):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(script_dir, 'Intensity calcs.xlsx')

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    intensity_lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            name = str(row[0]).strip()
            intensity_lookup[name] = {
                "description": row[1],
                "probability": row[2],
                "payout": row[3],
                "scale_min": row[4],
                "scale_max": row[5]
            }
    return intensity_lookup


def calculate_pay_split(total_pay, duration):
    possible_splits = []
    for m in range(0, (total_pay // duration) + 5, 5):
        total_retainer = m * duration
        bonus = total_pay - total_retainer
        if total_retainer >= 0.25 * total_pay and bonus >= 0.25 * total_pay:
            if bonus % 5 == 0:
                possible_splits.append((m, bonus))
    if possible_splits:
        return random.choice(possible_splits)
    m_raw = (total_pay // 2) // duration
    m_fallback = 5 * round(m_raw / 5)
    bonus_fallback = total_pay - (m_fallback * duration)
    return m_fallback, bonus_fallback


def calculate_scale_pv(intensity_name, intensity_data):
    stats = intensity_data.get(intensity_name.strip(), {})
    scale_min = stats.get('scale_min', 1.0)
    scale_max = stats.get('scale_max', 1.0)
    multiplier = random.uniform(scale_min, scale_max)
    raw_pv = 120 * multiplier
    return int(5 * round(raw_pv / 5))


def get_monthly_mission_count(intensity_name):
    intensity_rules = {
        "Very low": {"counts": [0, 1], "weights": [75, 25]},
        "Low": {"counts": [0, 1], "weights": [50, 50]},
        "Medium": {"counts": [0, 1, 2], "weights": [20, 60, 20]},
        "High": {"counts": [1, 2, 3], "weights": [20, 60, 20]},
        "Very high": {"counts": [2, 3, 4], "weights": [20, 60, 20]}
    }
    rule = intensity_rules.get(intensity_name.strip(), intensity_rules["Medium"])
    return random.choices(rule["counts"], weights=rule["weights"])[0]


def generate_mission_schedule(duration, intensity_name, campaign_type, campaign_data):
    schedule = []
    available_missions = campaign_data.get(campaign_type, {}).get('missions', ["Standard Combat"])
    last_mission = None
    for month in range(1, duration + 1):
        count = get_monthly_mission_count(intensity_name)
        month_missions = []
        for _ in range(count):
            valid_choices = [m for m in available_missions if m != last_mission]
            if not valid_choices:
                valid_choices = available_missions
            current_mission = random.choice(valid_choices)
            month_missions.append(current_mission)
            last_mission = current_mission
        schedule.append({"month": month, "count": count, "types": month_missions})
    return schedule


def parse_contract_parameters():
    file_path = 'contract_parameters.xlsx'
    if not os.path.exists(file_path):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(script_dir, 'contract_parameters.xlsx')
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    actors, salvage_terms = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]: actors.append(str(row[0]).strip())
        if row[1]: salvage_terms.append(str(row[1]).strip())
    return {"actors": actors, "salvage": salvage_terms}


def generate_random_campaign(campaign_data, intensity_data, contract_params):
    # 1. Pick random campaign and core details
    campaign_type = random.choice(list(campaign_data.keys()))
    details = campaign_data[campaign_type]

    intensity_name = random.choice(details.get('intensity', ["Standard"]))
    duration = random.choice(details.get('duration', [1]))

    # 2. Pick Employer, Opponent, and Salvage
    employer = random.choice(contract_params["actors"])
    remaining_actors = [a for a in contract_params["actors"] if a != employer]
    opponent = random.choice(remaining_actors)
    salvage = random.choice(contract_params["salvage"])

    # 3. Get intensity stats
    stats = intensity_data.get(intensity_name.strip(), {})
    intensity_description = stats.get('description', "No description available")
    base_monthly_payout = stats.get('payout', 0)

    # 4. Calculate Scale PV
    pv_value = calculate_scale_pv(intensity_name, intensity_data)

    # --- NEW: Calculate OpFor Size ---
    # Get the OpFor multiplier from the spreadsheet for this campaign
    opfor_multipliers = details.get('opfor', [1.0])
    # Filter out any non-numeric data that might have leaked in
    opfor_multipliers = [m for m in opfor_multipliers if isinstance(m, (int, float))]
    selected_opfor_mult = random.choice(opfor_multipliers) if opfor_multipliers else 1.0

    raw_opfor_size = pv_value * selected_opfor_mult

    # 1.1x bonus for 1-month duration
    if duration == 1:
        raw_opfor_size *= 1.1

    # Round to the nearest 5 for digestibility
    opfor_size = int(5 * round(raw_opfor_size / 5))

    # 5. Adjust Financials by PV and 20% Variance
    base_total_pay = duration * base_monthly_payout
    pv_adjusted_pay = (pv_value / 120) * base_total_pay
    variance_factor = random.uniform(0.8, 1.2)
    final_raw_pay = pv_adjusted_pay * variance_factor
    digestible_total_pay = int(5 * round(final_raw_pay / 5))

    monthly_retainer, bonus = calculate_pay_split(digestible_total_pay, duration)

    # 6. Generate the Mission Schedule
    mission_schedule = generate_mission_schedule(duration, intensity_name, campaign_type, campaign_data)

    # 7. Format outputs
    duration_text = f"{duration} month" if duration == 1 else f"{duration} months"

    output = [
        f"Employer: {employer}",
        f"Opponent: {opponent}",
        f"Campaign: {campaign_type}",
        f"Intensity: {intensity_name} ({intensity_description})",
        f"Duration: {duration_text}",
        f"Salvage: {salvage}",
        f"Pay (retainer): {monthly_retainer} per month",
        f"Pay (bonus): {bonus} one-time",
        f"Scale: Lance (minimum of {pv_value} PV)",
        f"OpFor size: ({opfor_size} PV)",
        "\nMission Schedule:"
    ]

    for item in mission_schedule:
        m_text = "mission" if item["count"] == 1 else "missions"
        line = f"Month {item['month']}: {item['count']} {m_text}"
        if item["count"] > 0:
            line += f" ({', '.join(item['types'])})"
        output.append(line)

    return "\n".join(output)


# --- EXECUTION ---
st.set_page_config(page_title="BattleTech Campaign Generator")
st.title("Campaign Generator")

if st.button("Generate New Campaign"):
    try:
        all_campaigns = parse_campaign_data()
        intensity_stats = parse_intensity_data()
        contract_params = parse_contract_parameters()

        random_setup = generate_random_campaign(all_campaigns, intensity_stats, contract_params)

        st.text_area("Result", value=random_setup, height=450)
    except Exception as e:
        st.error(f"An error occurred: {e}")